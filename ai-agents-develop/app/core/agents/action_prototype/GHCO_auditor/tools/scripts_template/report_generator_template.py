#!/usr/bin/env python3
"""
Standalone Report Generator Script

This script combines the entire report generation process into a single executable file.
It takes a list of ReportUnit data and mapping dictionaries as inputs and generates
an Excel report with summary and detailed information.

Usage:
    python standalone_report_generator.py
"""

import asyncio
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class EstimateUnit:
    """Data class for estimating yearly population."""

    time_delta: int
    number: int


@dataclass
class ReportUnit:
    """
    Data class representing a unit of report data containing account changes.
    """

    new_rows: pd.DataFrame
    delete_rows: pd.DataFrame
    update_rows: pd.DataFrame
    unique_columns: List[str]
    is_update_column: str
    business_unit: str
    period_start_date: datetime
    period_end_date: datetime


def estimate_yearly_population(data: List[EstimateUnit]) -> float:
    """
    Estimate yearly population based on time periods and frequencies.

    Args:
        data: List of EstimateUnit objects containing time periods and counts

    Returns:
        Estimated yearly population rounded to one decimal place

    Raises:
        ValueError: If any time_delta is zero or negative
    """
    # Validate input
    for unit in data:
        if unit.time_delta <= 0:
            raise ValueError(f"time_delta must be positive, got {unit.time_delta}")

    res = 0.0

    for unit in data:
        # Extrapolate to yearly rate based on the time period and frequency
        yearly_rate = (365.0 / unit.time_delta) * unit.number
        res += yearly_rate

    return round(res, 1)


def random_sample_rows(df: pd.DataFrame, sample: int) -> pd.DataFrame:
    """
    Randomly sample rows from a DataFrame.

    Args:
        df: DataFrame containing the rows to sample
        sample: Number of rows to sample

    Returns:
        DataFrame containing the sampled rows
    """
    df["Sample"] = ""
    if sample is None or sample <= 0 or df.empty:
        return df
    n = min(int(sample), len(df))
    sampled_rows = df.sample(n=n)
    df.loc[sampled_rows.index, "Sample"] = "x"
    return df


def map_frequency(value: float, frequency_mapping: Dict[str, str] = None) -> str:
    """
    Map estimated yearly population to frequency using dictionary lookup.

    Args:
        value: The estimated yearly population value
        frequency_mapping: Dictionary mapping population ranges to frequencies

    Returns:
        The corresponding frequency as a string
    """
    if frequency_mapping is None:
        # Default mapping
        if value > 260:
            return "Multiple Times Per Day"
        elif value >= 53:
            return "Daily"
        elif value >= 13:
            return "Weekly"
        elif value >= 5:
            return "Monthly"
        elif value >= 3:
            return "Quarterly"
        elif value == 2:
            return "Semi-Annual"
        elif value == 1:
            return "Annual"
        else:
            return "None"

    # Use custom mapping dictionary
    for range_key, frequency in frequency_mapping.items():
        if range_key == "> 260" and value > 260:
            return frequency
        elif range_key == "53-260" and 53 <= value <= 260:
            return frequency
        elif range_key == "13-52" and 13 <= value <= 52:
            return frequency
        elif range_key == "5-12" and 5 <= value <= 12:
            return frequency
        elif range_key == "3-4" and 3 <= value <= 4:
            return frequency
        elif range_key == "2" and value == 2:
            return frequency
        elif range_key == "1" and value == 1:
            return frequency

    return "None"


def map_sample_size(frequency: str, sample_size_mapping: Dict[str, int] = None) -> int:
    """
    Map frequency to sample size using dictionary lookup.

    Args:
        frequency: The frequency string
        sample_size_mapping: Dictionary mapping frequencies to sample sizes

    Returns:
        The corresponding sample size as an integer
    """
    if sample_size_mapping is None:
        # Default mapping
        default_mapping = {
            "Multiple Times Per Day": 20,
            "Daily": 15,
            "Weekly": 3,
            "Monthly": 1,
            "Quarterly": 1,
            "Semi-Annual": 1,
            "Annual": 1,
        }
        return default_mapping.get(frequency, 0)

    return sample_size_mapping.get(frequency, 0)


class ReportGenerator:
    """
    Manages the generation of Excel reports based on account change data.

    This class collects ReportUnit instances and creates Excel reports with
    summary and detailed information about account changes.
    """

    def __init__(
        self,
        frequency_mapping: Dict[str, str] = None,
        sample_size_mapping: Dict[str, int] = None,
    ):
        """
        Initialize the ReportGenerator with an empty list of report data.

        Args:
            frequency_mapping: Dictionary mapping population ranges to frequencies
            sample_size_mapping: Dictionary mapping frequencies to sample sizes
        """
        self.data: List[ReportUnit] = []
        self.frequency_mapping = frequency_mapping
        self.sample_size_mapping = sample_size_mapping
        self.add_sample_size = 0
        self.del_sample_size = 0

    def _preprocess_data(self, data: pd.DataFrame) -> pd.DataFrame:
        """
        Convert array-like objects in the DataFrame to pure strings.

        Args:
            data: A DataFrame containing the data to preprocess

        Returns:
            A DataFrame with array-like objects converted to strings
        """
        for column in data.columns:
            if data[column].apply(lambda x: isinstance(x, (list, tuple))).any():
                data[column] = data[column].apply(
                    lambda x: (
                        ", ".join(map(str, x)) if isinstance(x, (list, tuple)) else x
                    )
                )
        return data

    def add_result(self, result: ReportUnit):
        """
        Add a ReportUnit result to the report data.

        Args:
            result: A ReportUnit instance containing account change data

        The entries in the result are sorted by their unique columns before being added.
        """

        def _sort_by_unique_columns(df: pd.DataFrame, unique_cols: List[str]):
            df.sort_values(by=unique_cols, inplace=True)

        _sort_by_unique_columns(result.new_rows, result.unique_columns)
        _sort_by_unique_columns(result.delete_rows, result.unique_columns)
        _sort_by_unique_columns(result.update_rows, result.unique_columns)

        result.new_rows = self._preprocess_data(result.new_rows)
        result.delete_rows = self._preprocess_data(result.delete_rows)
        result.update_rows = self._preprocess_data(result.update_rows)

        self.data.append(result)
        self.data.sort(key=lambda x: x.business_unit)

    async def generate_report(self, file_path: str):
        """
        Generate an Excel report at the specified file path.

        Args:
            file_path: Path where the Excel file will be saved
        """
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet
        summary_sheet = workbook.create_sheet("Summary")
        added_updated_sheet = workbook.create_sheet("Added_Updated")
        deleted_data_sheet = workbook.create_sheet("Deleted")

        await self.write_to_summary(summary_sheet)
        self.write_to_added_updated(added_updated_sheet)
        self.write_to_deleted(deleted_data_sheet)

        workbook.save(file_path)

    async def write_to_summary(self, sheet: Worksheet):
        """
        Write summary information to the provided Excel sheet.

        Args:
            sheet: The openpyxl worksheet to write to
        """
        sheet.append(
            [
                "Business Unit",
                "Period Start Date",
                "Period End Date",
                "New Accounts",
                "Updated Accounts",
                "Deleted Accounts",
                "",
                "",
                "Test",
                "Estimated Yearly Population",
                "Frequency",
                "Sample Size",
            ]
        )

        for result in self.data:
            sheet.append(
                [
                    result.business_unit,
                    result.period_start_date.strftime("%Y-%m-%d"),
                    result.period_end_date.strftime("%Y-%m-%d"),
                    len(result.new_rows),
                    len(result.update_rows),
                    len(result.delete_rows),
                ]
            )

        sheet["I2"] = (
            "(AC2) User Provisioning and De-Provisioning - New / Added Account Provisioning"
        )
        sheet["I3"] = (
            "(AC2) User Provisioning and De-Provisioning - Account De-provisioning (Manual Termination)"
        )

        # Write estimated yearly population
        add_update_est = estimate_yearly_population(
            [
                EstimateUnit(
                    time_delta=(i.period_end_date - i.period_start_date).days,
                    number=len(i.new_rows) + len(i.update_rows),
                )
                for i in self.data
            ]
        )

        sheet["J2"] = add_update_est

        del_est = estimate_yearly_population(
            [
                EstimateUnit(
                    time_delta=(i.period_end_date - i.period_start_date).days,
                    number=len(i.delete_rows),
                )
                for i in self.data
            ]
        )
        sheet["J3"] = del_est

        add_feq = map_frequency(add_update_est, self.frequency_mapping)
        sheet["K2"] = add_feq

        del_feq = map_frequency(del_est, self.frequency_mapping)
        sheet["K3"] = del_feq

        add_sample_size = map_sample_size(add_feq, self.sample_size_mapping)
        sheet["L2"] = add_sample_size
        self.add_sample_size = add_sample_size

        del_sample_size = map_sample_size(del_feq, self.sample_size_mapping)
        sheet["L3"] = del_sample_size
        self.del_sample_size = del_sample_size

    def _get_unique_id(self, row: dict, unique_columns: List[str]) -> str:
        """
        Generate a unique ID string from a row based on the specified columns.

        Args:
            row: Dictionary representing a row of data
            unique_columns: List of column names to use for the unique ID

        Returns:
            A string representing the unique ID
        """
        return "_".join(str(row[col]) for col in unique_columns)

    def write_to_deleted(self, sheet: Worksheet):
        """Write deleted data to the Excel sheet."""
        for result in self.data:
            result.delete_rows["Business Unit"] = result.business_unit
            result.delete_rows["Unique ID"] = result.delete_rows.apply(
                lambda row: self._get_unique_id(row, result.unique_columns), axis=1
            )
            result.delete_rows["Unique ID Type"] = "_".join(result.unique_columns)
            result.delete_rows["Prior Roles"] = result.delete_rows.apply(
                lambda row: str(row.get(result.is_update_column)), axis=1
            )
            result.delete_rows["Current Roles"] = "-"

        full_df = pd.concat([i.delete_rows for i in self.data], ignore_index=True)
        sampled_full_df = random_sample_rows(full_df, self.del_sample_size)

        sheet.append(
            [
                "Sample",
                "Action",
                "Business Unit",
                "Unique ID",
                "Unique ID Type",
                "Prior Roles",
                "Current Roles",
            ]
        )

        for i, row in sampled_full_df.iterrows():
            sheet.append(
                [
                    row["Sample"],
                    "Deleted",
                    row["Business Unit"],
                    row["Unique ID"],
                    row["Unique ID Type"],
                    row["Prior Roles"],
                    row["Current Roles"],
                ]
            )

    def write_to_added_updated(self, sheet: Worksheet):
        """Write added and updated data to the Excel sheet."""
        for result in self.data:
            result.new_rows["Business Unit"] = result.business_unit
            result.new_rows["Action"] = "Added"
            result.new_rows["Unique ID"] = result.new_rows.apply(
                lambda row: self._get_unique_id(row, result.unique_columns), axis=1
            )
            result.new_rows["Unique ID Type"] = "_".join(result.unique_columns)
            result.new_rows["Prior Roles"] = "-"
            result.new_rows["Current Roles"] = result.new_rows.apply(
                lambda row: str(row.get(result.is_update_column)), axis=1
            )

        for result in self.data:
            result.update_rows["Business Unit"] = result.business_unit
            result.update_rows["Action"] = "Updated"
            result.update_rows["Unique ID"] = result.update_rows.apply(
                lambda row: self._get_unique_id(row, result.unique_columns), axis=1
            )
            result.update_rows["Unique ID Type"] = "_".join(result.unique_columns)
            result.update_rows["Prior Roles"] = result.update_rows.apply(
                lambda row: str(row.get(f"{result.is_update_column}_old")), axis=1
            )
            result.update_rows["Current Roles"] = result.update_rows.apply(
                lambda row: str(row.get(f"{result.is_update_column}_new")), axis=1
            )

        full_new_df = pd.concat([i.new_rows for i in self.data], ignore_index=True)
        full_update_df = pd.concat(
            [i.update_rows for i in self.data], ignore_index=True
        )
        full_df = pd.concat([full_new_df, full_update_df], ignore_index=True)

        sampled_full_df = random_sample_rows(full_df, self.add_sample_size)

        sheet.append(
            [
                "Sample",
                "Action",
                "Business Unit",
                "Unique ID",
                "Unique ID Type",
                "Prior Roles",
                "Current Roles",
            ]
        )

        for i, row in sampled_full_df.iterrows():
            sheet.append(
                [
                    row["Sample"],
                    row["Action"],
                    row["Business Unit"],
                    row["Unique ID"],
                    row["Unique ID Type"],
                    row["Prior Roles"],
                    row["Current Roles"],
                ]
            )


async def generate_report_standalone(
    report_units: List[ReportUnit],
    output_file_path: str = "./generated_report.xlsx",
    frequency_mapping: Dict[str, str] = None,
    sample_size_mapping: Dict[str, int] = None,
) -> None:
    """
    Generate a report using the provided data.

    Args:
        report_units: List of ReportUnit objects containing the data to process
        output_file_path: Path where the Excel report will be saved
        frequency_mapping: Dictionary mapping population ranges to frequencies
        sample_size_mapping: Dictionary mapping frequencies to sample sizes
    """
    # Initialize report generator with custom mappings
    report_generator = ReportGenerator(
        frequency_mapping=frequency_mapping, sample_size_mapping=sample_size_mapping
    )

    # Process each report unit
    for unit in report_units:
        report_generator.add_result(unit)

    # Generate the Excel report
    await report_generator.generate_report(output_file_path)


# Default mapping dictionaries for reference
DEFAULT_FREQUENCY_MAPPING = {
    "> 260": "Multiple Times Per Day",
    "53-260": "Daily",
    "13-52": "Weekly",
    "5-12": "Monthly",
    "3-4": "Quarterly",
    "2": "Semi-Annual",
    "1": "Annual",
}

DEFAULT_SAMPLE_SIZE_MAPPING = {
    "Multiple Times Per Day": 20,
    "Daily": 15,
    "Weekly": 3,
    "Monthly": 1,
    "Quarterly": 1,
    "Semi-Annual": 1,
    "Annual": 1,
}


def create_sample_report_unit() -> ReportUnit:
    """
    Create a sample ReportUnit for testing purposes.

    Returns:
        A sample ReportUnit with test data
    """
    # Create sample DataFrames
    new_rows = pd.DataFrame(
        {
            "Username": ["user1", "user2"],
            "Groups": ["admin", "user"],
            "Email": ["user1@example.com", "user2@example.com"],
        }
    )

    delete_rows = pd.DataFrame(
        {"Username": ["user3"], "Groups": ["guest"], "Email": ["user3@example.com"]}
    )

    update_rows = pd.DataFrame(
        {
            "Username": ["user4"],
            "Groups_old": ["user"],
            "Groups_new": ["admin"],
            "Email": ["user4@example.com"],
        }
    )

    return ReportUnit(
        new_rows=new_rows,
        delete_rows=delete_rows,
        update_rows=update_rows,
        unique_columns=["Username"],
        is_update_column="Groups",
        business_unit="SAMPLE",
        period_start_date=datetime(2024, 1, 1),
        period_end_date=datetime(2024, 12, 31),
    )


if __name__ == "__main__":
    # Example usage with sample data
    sample_unit = create_sample_report_unit()
    report_units = [sample_unit]

    # Generate sample report
    asyncio.run(
        generate_report_standalone(
            report_units=report_units, output_file_path="./sample_report.xlsx"
        )
    )
